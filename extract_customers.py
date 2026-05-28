import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

XLSX = r'C:\Users\M\Desktop\github for local\adrenaline-last\ADRENALINE 1-6-2026.xlsx'

wb = openpyxl.load_workbook(XLSX, data_only=True)

def parse_dates(s, year=2026):
    if not s: return None, None
    s = str(s).strip()
    m = re.match(r'(\d{1,2})\s*[-/]\s*(\d{1,2})\s+(?:END|to|TO)\s+(\d{1,2})\s*[-/]\s*(\d{1,2})', s, re.I)
    if m:
        sd, sm, ed, em = m.groups()
        start_year = year
        # If start is Nov/Dec and we're talking about 2026, year is 2025
        # If end_month < start_month, end goes to next year
        if int(sm) > 6:  # If start in second half, probably continuing from previous year
            start_year = year - 1 if int(sm) >= 11 else year
        end_year = start_year if int(em) >= int(sm) else start_year + 1
        try:
            return f"{start_year}-{int(sm):02d}-{int(sd):02d}", f"{end_year}-{int(em):02d}-{int(ed):02d}"
        except: return None, None
    # Single date "1-6-2026" or "31-5"
    m2 = re.match(r'(\d{1,2})\s*[-/]\s*(\d{1,2})$', s)
    if m2:
        return f"{year}-{int(m2.group(2)):02d}-{int(m2.group(1)):02d}", None
    return None, None

def parse_package(s):
    if not s: return None, None, None
    s = str(s)
    m_meals = re.search(r'(\d+)\s*M\b', s, re.I)
    m_snacks = re.search(r'(\d+)\s*S\b', s, re.I)
    m_weeks = re.search(r'(\d+)\s*W', s, re.I)
    return (
        int(m_meals.group(1)) if m_meals else None,
        int(m_snacks.group(1)) if m_snacks else None,
        int(m_weeks.group(1)) if m_weeks else None,
    )

def split_program(raw):
    """'FITNESS /NO TOMATO,ONIONS' -> program='FITNESS', avoid='NO TOMATO,ONIONS'"""
    if not raw: return '', ''
    s = str(raw).strip()
    # split on first /
    parts = re.split(r'\s*/\s*', s, maxsplit=1)
    program = parts[0].strip().upper()
    # Normalize program
    PROGS = {'BULK', 'FITNESS', 'DIET', 'FITNTESS', 'FITNEESS', 'CUSTOMIZED'}
    if program not in PROGS:
        # Maybe combined, try first word
        first = program.split()[0] if program else ''
        if first in PROGS:
            program = first
    # Normalize typos
    program = program.replace('FITNTESS', 'FITNESS').replace('FITNEESS', 'FITNESS')
    notes = parts[1].strip() if len(parts) > 1 else ''
    return program, notes

def norm_phone(x):
    if not x: return ''
    s = re.sub(r'\D', '', str(x))
    # Qatar normalization
    if s.startswith('974'): s = s[3:]
    return s

def norm_dt(x):
    if not x: return None
    s = str(x).strip().upper()
    if 'EVEN' in s or 'NIGHT' in s: return 'EVENING'
    if 'MORN' in s: return 'MORNING'
    return None

customers = []
dups = set()  # (phone, name) pair to avoid exact duplicates only

# ==== SHEET ====
ws = wb['SHEET']
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or not row[1]: continue
    phone = norm_phone(row[1])
    name = (row[2] or '').strip() if row[2] else ''
    if not phone or not name: continue
    key = (phone, name.upper())  # allow same phone with different names (family)
    if key in dups: continue
    dups.add(key)

    start, end = parse_dates(row[3])
    program, avoid_notes = split_program(row[4])
    pkg_label = (str(row[5]).strip() if row[5] else '')
    meals, snacks, weeks = parse_package(pkg_label)

    # Find delivery time (try cols 11, 12, 13)
    delivery = None
    for idx in (11, 12, 13):
        if len(row) > idx and row[idx]:
            d = norm_dt(row[idx])
            if d:
                delivery = d
                break
    if not delivery:
        delivery = 'MORNING'

    customers.append({
        'phone': phone,
        'fullName': name,
        'startDate': start or '2026-05-01',
        'endDate': end or '2026-05-31',
        'program': program or 'STANDARD',
        'packageLabel': pkg_label,
        'mealsPerDay': meals,
        'snacksPerDay': snacks,
        'durationWeeks': weeks,
        'totalMealsPerDay': (meals or 0) + (snacks or 0) if (meals or snacks) else None,
        'deliveryTime': delivery,
        'avoid': avoid_notes or None,
        'notes': avoid_notes or None,
        'isActive': True,
        '_source': 'SHEET',
    })

# ==== CUTOMIZED ====
ws = wb['CUTOMIZED']
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or not row[1]: continue
    phone = norm_phone(row[1])
    name = (row[2] or '').strip() if row[2] else ''
    key = (phone, name.upper())
    if not phone or not name or key in dups: continue
    dups.add(key)

    start, end = parse_dates(row[3])
    info = (str(row[4]).strip() if row[4] else '')
    meals, snacks, weeks = parse_package(info)

    delivery = None
    for idx in (11, 12, 13):
        if len(row) > idx and row[idx]:
            d = norm_dt(row[idx])
            if d:
                delivery = d
                break
    if not delivery:
        delivery = 'MORNING'

    customers.append({
        'phone': phone,
        'fullName': name,
        'startDate': start or '2026-05-01',
        'endDate': end or '2026-05-31',
        'program': 'CUSTOMIZED',
        'packageLabel': info,
        'mealsPerDay': meals,
        'snacksPerDay': snacks,
        'durationWeeks': weeks,
        'totalMealsPerDay': (meals or 0) + (snacks or 0) if (meals or snacks) else None,
        'deliveryTime': delivery,
        'allergies': info,
        'isActive': True,
        '_source': 'CUTOMIZED',
    })

# Clean rows - remove None and _source for JSON
clean = []
for c in customers:
    cleaned = {k: v for k, v in c.items() if v is not None and v != '' and not k.startswith('_')}
    cleaned.setdefault('isActive', True)
    clean.append(cleaned)

print(f'Total: {len(clean)}')
prog_counts = {}
for c in clean:
    p = c.get('program', '?')
    prog_counts[p] = prog_counts.get(p, 0) + 1
print(f'By program: {prog_counts}')

dt_counts = {}
for c in clean:
    p = c.get('deliveryTime', '?')
    dt_counts[p] = dt_counts.get(p, 0) + 1
print(f'By delivery: {dt_counts}')

print('\nSample 5:')
for c in clean[:5]:
    print(f"  {c['phone']:>10} | {c['fullName'][:25]:25} | {c['startDate']} -> {c['endDate']} | {c.get('program','?'):10} | {c['deliveryTime']} | M={c.get('mealsPerDay')} S={c.get('snacksPerDay')}")

with open(r'C:\Users\M\Desktop\github for local\adrenaline-last\customers-import.json', 'w', encoding='utf-8') as f:
    json.dump(clean, f, ensure_ascii=False, indent=2)
print('\n[OK] Saved customers-import.json')
